"""
Export an .xlsx worksheet to CSV.
Usage:
    python scripts/export_xlsx_to_csv.py "C:\path\to\file.xlsx" "output.csv"
"""
import sys
import os
import csv
from openpyxl import load_workbook


def export(path, outpath):
    if not os.path.exists(path):
        print('File not found:', path)
        return False
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    os.makedirs(os.path.dirname(outpath), exist_ok=True)
    with open(outpath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(['' if c is None else str(c) for c in row])
    wb.close()
    return True


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Usage: python scripts/export_xlsx_to_csv.py input.xlsx output.csv')
        sys.exit(1)
    input_path = sys.argv[1]
    out_path = sys.argv[2]
    ok = export(input_path, out_path)
    if ok:
        print('Exported to', out_path)
    else:
        sys.exit(2)
