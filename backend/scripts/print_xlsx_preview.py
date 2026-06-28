"""
Print headers and preview rows from an .xlsx file.
Usage:
    python scripts/print_xlsx_preview.py "C:\path\to\file.xlsx" [--rows N]
"""
import sys
import os
import argparse
from openpyxl import load_workbook


def preview(path, rows=20):
    if not os.path.exists(path):
        print('File not found:', path)
        return
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        print('Empty workbook')
        return
    print('Headers:')
    print([str(h) for h in header])
    print('\nPreview rows:')
    count = 0
    for row in it:
        print([str(c) for c in row])
        count += 1
        if count >= rows:
            break
    wb.close()


if __name__ == '__main__':
    p = argparse.ArgumentParser()
    p.add_argument('path')
    p.add_argument('--rows', type=int, default=50)
    args = p.parse_args()
    preview(args.path, rows=args.rows)
