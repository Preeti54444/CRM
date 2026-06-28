"""
Import Excel preserving exact lead history and timeline events.

This script stores the original row in `lead_import_raw` and then imports or updates
leads in the CRM while preserving historical timestamps.

Features:
- creates `lead_import_raw` if missing
- normalizes lead contact, company, source, status, and location
- matches existing leads by mobile or email
- preserves earliest created_at and latest updated_at values
- inserts timeline events for the import, calls, follow-ups, and final outcome
- maps Sales Executive to existing users when possible

Usage:
    python scripts/import_preserve_exact_excel.py "C:\path\to\lead journey.xlsx"
"""
from __future__ import annotations
import importlib
import json
import os
import pkgutil
import re
import sys
import uuid
from datetime import datetime
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func, text

# ensure app package importable
sys.path.insert(0, os.getcwd())
from app.database import SessionLocal, engine
from app.models.user import User
from app.services.user_service import get_user_by_email


EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def normalize_header(h: Any) -> str:
    if h is None:
        return ''
    return str(h).strip().lower()


def clean_string(raw: Any) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    return s if s else None


def parse_date_cell(val: Any) -> datetime | None:
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%d %b %Y', '%d %B %Y'):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    try:
        from dateutil.parser import parse

        return parse(s)
    except Exception:
        return None


def normalize_mobile(raw: Any) -> str | None:
    if raw is None:
        return None
    s = re.sub(r'[^0-9]', '', str(raw))
    if len(s) == 0:
        return None
    if len(s) > 10:
        s = s[-10:]
    return s


def parse_email(raw: Any) -> str | None:
    s = clean_string(raw)
    if not s:
        return None
    if EMAIL_RE.match(s.lower()):
        return s.lower()
    return None


def parse_location(raw: Any) -> tuple[str | None, str | None]:
    if raw is None:
        return None, None
    s = str(raw).strip()
    if not s:
        return None, None
    parts = [p.strip() for p in re.split(r'[\n\r,]|\s{2,}', s) if p.strip()]
    if len(parts) >= 2:
        return parts[-1], parts[0]
    return s, None


def parse_funding_amount(raw: Any) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().replace(',', '').replace('₹', '').replace('$', '').replace('Rs.', '').strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def truncate(s: Any, n: int) -> Any:
    if s is None:
        return None
    s2 = str(s)
    if len(s2) <= n:
        return s2
    return s2[:n]


def create_raw_table() -> None:
    with engine.begin() as conn:
        conn.execute(
            text(
                '''
            CREATE TABLE IF NOT EXISTS lead_import_raw (
                id SERIAL PRIMARY KEY,
                source_file TEXT,
                row_num INT,
                data JSONB,
                date_of_entry TIMESTAMPTZ,
                lead_id INT
            )
            '''
            )
        )


def get_header_index(headers: list[str]) -> dict[str, int]:
    result = {}
    for index, header in enumerate(headers):
        key = normalize_header(header)
        if key:
            result[key] = index
    return result


def get_row_value(row: tuple[Any, ...], headers: list[str], header_index: dict[str, int], *keys: str) -> Any:
    for key in keys:
        idx = header_index.get(normalize_header(key))
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def find_assignee_id(db, raw_sales_exec: Any) -> str | None:
    sales_exec = clean_string(raw_sales_exec)
    if not sales_exec:
        return None
    if '@' in sales_exec:
        user = get_user_by_email(db, sales_exec)
        if user:
            return str(user.id)
    normalized = sales_exec.lower()
    return_value = db.query(User).filter(func.lower(User.full_name) == normalized).first()
    if return_value:
        return str(return_value.id)
    return None


def insert_timeline_event(conn, lead_id: int, event_type: str, description: str | None, metadata: str | None, created_by: str | None, created_at: datetime | None) -> None:
    if created_at is None:
        created_at = datetime.utcnow()
    conn.execute(
        text(
            '''
        INSERT INTO timeline_events (id, lead_id, event_type, description, event_metadata, created_by, created_at)
        VALUES (:id, :lead_id, :event_type, :description, :event_metadata, :created_by, :created_at)
        '''
        ),
        {
            'id': uuid.uuid4(),
            'lead_id': lead_id,
            'event_type': event_type,
            'description': description,
            'event_metadata': metadata,
            'created_by': created_by,
            'created_at': created_at,
        },
    )


def insert_followup(conn, lead_id: int, assigned_to: str | None, followup_date: datetime, notes: str | None, next_followup_date: datetime | None, created_by: str | None, created_at: datetime) -> None:
    conn.execute(
        text(
            '''
        INSERT INTO followups (id, lead_id, assigned_to, followup_date, followup_type, notes, next_followup_date, status, created_by, created_at, updated_at)
        VALUES (:id, :lead_id, :assigned_to, :followup_date, :followup_type, :notes, :next_followup_date, :status, :created_by, :created_at, :updated_at)
        '''
        ),
        {
            'id': uuid.uuid4(),
            'lead_id': lead_id,
            'assigned_to': assigned_to,
            'followup_date': followup_date,
            'followup_type': 'Imported Follow-Up',
            'notes': notes,
            'next_followup_date': next_followup_date,
            'status': 'scheduled',
            'created_by': created_by,
            'created_at': created_at,
            'updated_at': created_at,
        },
    )


def make_serializable(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if value is None:
        return None
    return str(value)


def import_file(path: str, source_file_name: str | None = None, sheet_name: str | None = None) -> None:
    if not os.path.exists(path):
        print('File not found:', path)
        return
    source_file_name = source_file_name or os.path.basename(path)
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    iterator = ws.iter_rows(values_only=True)
    try:
        raw_headers = [h if h is not None else f'Column {i + 1}' for i, h in enumerate(next(iterator))]
    except StopIteration:
        print('Empty workbook')
        wb.close()
        return

    headers = get_header_index(raw_headers)
    create_raw_table()

    db = SessionLocal()
    created = 0
    updated = 0
    raw_saved = 0
    exported_rows = 0
    skipped = 0
    duplicate_count = 0

    try:
        for rownum, row in enumerate(iterator, start=2):
            exported_rows += 1
            row_data = {raw_headers[i]: row[i] if i < len(row) else None for i in range(len(raw_headers))}
            date_of_entry = parse_date_cell(get_row_value(row, raw_headers, headers, 'Date of Entry', 'Date of Entry '))
            first_call_date = parse_date_cell(get_row_value(row, raw_headers, headers, 'Date of First Call'))
            followup_date = parse_date_cell(get_row_value(row, raw_headers, headers, 'Next Follow-up Date'))
            sales_exec = get_row_value(row, raw_headers, headers, 'Sales Executive', 'Sales Executive ')
            lead_source = clean_string(get_row_value(row, raw_headers, headers, 'Lead Source', '  Lead Source  '))
            company_name = clean_string(get_row_value(row, raw_headers, headers, 'Customer Company Name', '  Customer Company Name  '))
            contact_name = clean_string(get_row_value(row, raw_headers, headers, 'Contact Person Name', 'Contact Person Name  '))
            contact_number = normalize_mobile(get_row_value(row, raw_headers, headers, 'Contact Number', 'Contact Number  ', 'Contact No', 'Contact no'))
            email_id = parse_email(get_row_value(row, raw_headers, headers, 'Email ID', 'Email ID  '))
            email_address = parse_email(get_row_value(row, raw_headers, headers, 'Email Address', 'Email Address '))
            location = get_row_value(row, raw_headers, headers, 'Location (City, State)', 'Location (City, State)  ')
            current_status = clean_string(get_row_value(row, raw_headers, headers, 'Current Status of Lead', 'Current Status of Lead  '))
            proposal_shared = clean_string(get_row_value(row, raw_headers, headers, 'Proposal/Document Shared?  Yes / No  (If Yes, Mention Date)  '))
            purpose_of_call = clean_string(get_row_value(row, raw_headers, headers, 'Purpose of Call', 'Purpose of Call  '))
            product_discussed = clean_string(get_row_value(row, raw_headers, headers, 'Product/Service Discussed', 'Product/Service Discussed  '))
            call_outcome = clean_string(get_row_value(row, raw_headers, headers, 'Call Outcome', 'Call Outcome  '))
            final_outcome = clean_string(get_row_value(row, raw_headers, headers, 'Final Outcome', 'Final Outcome  '))
            deal_value = parse_funding_amount(get_row_value(row, raw_headers, headers, 'Deal Value (If Closed)', 'Deal Value (If Closed)  '))
            learning_challenge = clean_string(get_row_value(row, raw_headers, headers, 'Learning / Challenge Faced', 'Learning / Challenge Faced  '))

            city, state = parse_location(location)
            assigned_to = find_assignee_id(db, sales_exec)

            email = email_id or email_address
            company_email = email_address if email_id and email_address != email_id else None

            lead_name = contact_name or company_name or email or contact_number or f'Imported Lead {rownum}'
            remarks_parts = []
            if proposal_shared:
                remarks_parts.append(f'Proposal shared: {proposal_shared}')
            if call_outcome:
                remarks_parts.append(f'Call outcome: {call_outcome}')
            if final_outcome:
                remarks_parts.append(f'Final outcome: {final_outcome}')
            if learning_challenge:
                remarks_parts.append(f'Learning/challenge: {learning_challenge}')
            if purpose_of_call:
                remarks_parts.append(f'Call purpose: {purpose_of_call}')
            if product_discussed:
                remarks_parts.append(f'Product discussed: {product_discussed}')
            remarks = ' | '.join(remarks_parts) if remarks_parts else None

            # enforce DB column size limits
            lead_name = truncate(lead_name, 255)
            company_name = truncate(company_name, 255)
            lead_status = truncate(current_status or final_outcome or 'New', 100)
            product_type = truncate(product_discussed or None, 100)
            lead_source = truncate(lead_source or None, 100)
            remarks = truncate(remarks, 1000)
            funding_amount = deal_value

            raw_serializable = {k: make_serializable(v) for k, v in row_data.items()}

            with engine.begin() as conn:
                conn.execute(
                    text(
                        'INSERT INTO lead_import_raw(source_file, row_num, data, date_of_entry) VALUES (:sf, :rn, :data, :doe)'
                    ),
                    {'sf': source_file_name, 'rn': rownum, 'data': json.dumps(raw_serializable), 'doe': date_of_entry},
                )
                raw_saved += 1

            existing_lead = None
            with engine.connect() as conn:
                if contact_number:
                    existing_lead = conn.execute(
                        text('SELECT id, lead_name, company_name, email, mobile, created_at, updated_at FROM leads WHERE mobile = :m LIMIT 1'),
                        {'m': contact_number},
                    ).fetchone()
                if not existing_lead and email:
                    existing_lead = conn.execute(
                        text('SELECT id, lead_name, company_name, email, mobile, created_at, updated_at FROM leads WHERE email = :e LIMIT 1'),
                        {'e': email},
                    ).fetchone()

            if existing_lead:
                lead_id = existing_lead[0]
                updates = {}
                if not existing_lead[1] and lead_name:
                    updates['lead_name'] = lead_name
                if not existing_lead[2] and company_name:
                    updates['company_name'] = company_name
                if not existing_lead[3] and email:
                    updates['email'] = email
                if not existing_lead[4] and contact_number:
                    updates['mobile'] = contact_number
                if company_email and not existing_lead[3]:
                    updates['company_email'] = company_email
                if product_type:
                    updates['product_type'] = product_type
                if lead_source:
                    updates['lead_source'] = lead_source
                if lead_status and existing_lead[5] is not None and existing_lead[6] is not None:
                    updates['lead_status'] = lead_status
                if assigned_to:
                    updates['assigned_to'] = assigned_to
                if remarks:
                    updates['remarks'] = remarks
                if date_of_entry:
                    updates['created_at'] = date_of_entry
                    if existing_lead[6] is None or date_of_entry > existing_lead[6]:
                        updates['updated_at'] = date_of_entry
                if updates:
                    set_clause = ', '.join(f'{k} = :{k}' for k in updates)
                    params = updates.copy()
                    params['id'] = lead_id
                    with engine.begin() as conn:
                        conn.execute(text(f'UPDATE leads SET {set_clause} WHERE id = :id'), params)
                    updated += 1
                duplicate_count += 1
            else:
                insert_cols = [
                    'lead_name',
                    'company_name',
                    'mobile',
                    'alternate_mobile',
                    'email',
                    'company_email',
                    'city',
                    'state',
                    'product_type',
                    'funding_amount',
                    'lead_source',
                    'lead_status',
                    'assigned_to',
                    'created_by',
                    'remarks',
                ]
                values = {col: None for col in insert_cols}
                values.update(
                    {
                        'lead_name': lead_name,
                        'company_name': company_name,
                        'mobile': contact_number,
                        'alternate_mobile': None,
                        'email': email,
                        'company_email': company_email,
                        'city': city,
                        'state': state,
                        'product_type': product_type,
                        'funding_amount': funding_amount,
                        'lead_source': lead_source,
                        'lead_status': lead_status,
                        'assigned_to': assigned_to,
                        'created_by': assigned_to,
                        'remarks': remarks,
                    }
                )
                ts = date_of_entry or first_call_date or followup_date or datetime.utcnow()
                values['created_at'] = ts
                values['updated_at'] = ts
                insert_cols.extend(['created_at', 'updated_at'])
                cols_sql = ', '.join(insert_cols)
                params_sql = ', '.join(f':{c}' for c in insert_cols)
                with engine.begin() as conn:
                    res = conn.execute(text(f'INSERT INTO leads ({cols_sql}) VALUES ({params_sql}) RETURNING id'), values)
                    lead_id = res.fetchone()[0]
                created += 1

            with engine.begin() as conn:
                insert_timeline_event(
                    conn,
                    lead_id,
                    'Lead Imported',
                    f'Imported row {rownum} from {source_file_name}.',
                    json.dumps(raw_serializable),
                    assigned_to,
                    date_of_entry or datetime.utcnow(),
                )
                if first_call_date:
                    insert_timeline_event(
                        conn,
                        lead_id,
                        'Call Logged',
                        'First call logged. ' + ' '.join(
                            part for part in [purpose_of_call, product_discussed, call_outcome] if part
                        ),
                        None,
                        assigned_to,
                        first_call_date,
                    )
                if followup_date:
                    insert_timeline_event(
                        conn,
                        lead_id,
                        'Follow-Up Scheduled',
                        f'Follow-up scheduled for {followup_date.isoformat()}.',
                        None,
                        assigned_to,
                        followup_date,
                    )
                    insert_followup(
                        conn,
                        lead_id,
                        assigned_to,
                        followup_date,
                        'Imported follow-up from Excel.',
                        followup_date,
                        assigned_to,
                        date_of_entry or datetime.utcnow(),
                    )
                if final_outcome or deal_value or learning_challenge:
                    description_parts = []
                    if final_outcome:
                        description_parts.append(f'Final outcome: {final_outcome}')
                    if deal_value is not None:
                        description_parts.append(f'Deal value: {deal_value}')
                    if learning_challenge:
                        description_parts.append(f'Learning/challenge: {learning_challenge}')
                    insert_timeline_event(
                        conn,
                        lead_id,
                        'Final Outcome Recorded',
                        ' '.join(description_parts),
                        None,
                        assigned_to,
                        date_of_entry or datetime.utcnow(),
                    )

            with engine.begin() as conn:
                conn.execute(
                    text('UPDATE lead_import_raw SET lead_id = :lid WHERE source_file = :sf AND row_num = :rn'),
                    {'lid': lead_id, 'sf': source_file_name, 'rn': rownum},
                )

        print(
            f'Import finished. Rows={exported_rows}, Raw saved={raw_saved}, Created={created}, Updated={updated}, Duplicates={duplicate_count}, Skipped={skipped}'
        )
    finally:
        db.close()
        wb.close()


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python scripts/import_preserve_exact_excel.py <path-to-xlsx> [sheet-name]')
        sys.exit(1)
    path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) >= 3 else None
    import_file(path, sheet_name=sheet)
