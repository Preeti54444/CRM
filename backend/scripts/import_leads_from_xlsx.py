"""
Import leads from an Excel (.xlsx) file into the application's database.
Usage:
    python scripts/import_leads_from_xlsx.py "C:\path\to\lead journey.xlsx"

The script maps columns in the spreadsheet to the Lead fields used by the backend.
Expected header names (case-insensitive):
  lead_name, company_name, mobile, alternate_mobile, email, company_email,
  city, state, product_type, funding_amount, lead_source, lead_status, remarks

If a header is not present the field is left as None.
"""
from __future__ import annotations
import sys
import os
from typing import Dict, Any

from openpyxl import load_workbook

from app.database import SessionLocal
from app.schemas.lead import LeadCreate
from app.services.lead_service import create_lead


# Mapping of common header names to Lead fields
HEADER_MAP = {
    'lead_name': 'lead_name',
    'name': 'lead_name',
    'full_name': 'lead_name',
    'company_name': 'company_name',
    'mobile': 'mobile',
    'phone': 'mobile',
    'alternate_mobile': 'alternate_mobile',
    'mobile_alternate': 'alternate_mobile',
    'email': 'email',
    'company_email': 'company_email',
    'city': 'city',
    'state': 'state',
    'product_type': 'product_type',
    'loan_type': 'product_type',
    'funding_amount': 'funding_amount',
    'loan_amount': 'funding_amount',
    'lead_source': 'lead_source',
    'lead_status': 'lead_status',
    'remarks': 'remarks',
}


def normalize_header(h: Any) -> str:
    if h is None:
        return ''
    return str(h).strip().lower()


def row_to_lead_dict(headers, row) -> Dict[str, Any]:
    data: Dict[str, Any] = {}
    for idx, cell in enumerate(row):
        if idx >= len(headers):
            continue
        h = headers[idx]
        if not h:
            continue
        key = HEADER_MAP.get(h)
        if not key:
            continue
        val = cell.value
        if key == 'funding_amount' and val is not None:
            try:
                val = float(val)
            except Exception:
                try:
                    # remove commas/currency
                    val = float(str(val).replace(',', '').replace('₹', '').strip())
                except Exception:
                    val = None
        if isinstance(val, str):
            val = val.strip()
            if val == '':
                val = None
        data[key] = val
    return data


def import_file(path: str):
    if not os.path.exists(path):
        print('File not found:', path)
        return

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active

    iterator = ws.iter_rows(values_only=False)
    try:
        header_row = next(iterator)
    except StopIteration:
        print('Empty workbook')
        return

    headers = [normalize_header(c.value) for c in header_row]

    db = SessionLocal()
    created = 0
    skipped = 0
    try:
        for row in iterator:
            lead_data = row_to_lead_dict(headers, row)
            # Skip rows with no lead_name and no mobile
            if not lead_data.get('lead_name') and not lead_data.get('mobile'):
                skipped += 1
                continue
            try:
                lead_in = LeadCreate(**lead_data)
            except Exception as e:
                print('Validation failed for row, skipping:', e)
                skipped += 1
                continue
            created_lead = create_lead(db, lead_in)
            print(f"Inserted lead id={created_lead.id} name={created_lead.lead_name}")
            created += 1
    finally:
        db.close()
        wb.close()

    print(f"Import finished. Created={created} Skipped={skipped}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python scripts/import_leads_from_xlsx.py <path-to-xlsx>')
        sys.exit(1)
    path = sys.argv[1]
    import_file(path)
