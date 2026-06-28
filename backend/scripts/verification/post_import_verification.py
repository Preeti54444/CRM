import sys
from pathlib import Path
import json

repo_root = Path(__file__).resolve().parents[2]  # points to backend
sys.path.insert(0, str(repo_root))

from app import database, config
from sqlalchemy import text
import openpyxl
import requests
import time


def db_count(table: str) -> int:
    with database.engine.connect() as conn:
        r = conn.execute(text(f"SELECT COUNT(*) FROM {table}"))
        return int(r.scalar() or 0)


def excel_row_count(xlsx_path: str, sheet_name: str) -> int:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet {sheet_name} not found in workbook")
    ws = wb[sheet_name]
    # subtract header row
    return max(0, ws.max_row - 1)


def sample_lead_api(base_url: str, lead_id: int, token: str | None = None):
    url = f"{base_url}/leads/{lead_id}"
    headers = {}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    r = requests.get(url, headers=headers, timeout=10)
    return r.status_code, r.json() if r.status_code == 200 else r.text


def get_token(base_url: str, email: str, password: str) -> str | None:
    try:
        r = requests.post(f"{base_url}/auth/login", json={"email": email, "password": password}, timeout=10)
        if r.status_code == 200:
            return r.json().get("access_token")
        return None
    except Exception:
        return None


def main():
    # Config
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("../lead journey.xlsx").resolve()
    sheet = sys.argv[2] if len(sys.argv) > 2 else "Leadzz Journey FS"
    base_url = config.settings.frontend_url if config.settings.api_base is None else config.settings.api_base
    if base_url.endswith("/"):
        base_url = base_url[:-1]

    out = {}
    out["xlsx"] = str(xlsx)
    out["sheet"] = sheet

    print("Counting Excel rows...")
    out["excel_rows"] = excel_row_count(str(xlsx), sheet)
    print("Excel rows (excluding header):", out["excel_rows"])

    print("Counting DB leads...")
    out["db_leads"] = db_count("leads")
    print("DB leads:", out["db_leads"])

    print("Counting timeline events...")
    out["db_timeline_events"] = db_count("timeline_events")
    print("DB timeline events:", out["db_timeline_events"])

    # sample API check for a few lead ids (first 10)
    sample_ids = []
    with database.engine.connect() as conn:
        rows = conn.execute(text("SELECT id FROM leads ORDER BY id DESC LIMIT 10"))
        sample_ids = [r[0] for r in rows]

    out["api_samples"] = {}
    for lid in sample_ids:
        try:
            code, body = sample_lead_api(base_url, lid)
        except Exception as e:
            code, body = 0, str(e)
        out["api_samples"][str(lid)] = {"status": code, "body_preview": body if isinstance(body, dict) else str(body)[:400]}

    # attempt authenticated API checks
    admin_token = get_token(base_url, "shree.rathod@fundingsathi.in", "shree.admin@2026")
    out["admin_token_present"] = bool(admin_token)
    if admin_token:
        try:
            r = requests.get(f"{base_url}/leads?limit=5", headers={"Authorization": f"Bearer {admin_token}"}, timeout=10)
            out["admin_leads_list_status"] = r.status_code
            out["admin_leads_list_preview"] = r.json() if r.status_code == 200 else r.text[:400]
        except Exception as e:
            out["admin_leads_list_error"] = str(e)
        # additional admin checks: search, filter, timeline, reports
        try:
            r_search = requests.get(f"{base_url}/leads?search=Vaibhav&limit=5", headers={"Authorization": f"Bearer {admin_token}"}, timeout=10)
            out["admin_search_status"] = r_search.status_code
            out["admin_search_preview"] = r_search.json() if r_search.status_code == 200 else r_search.text[:400]
        except Exception as e:
            out["admin_search_error"] = str(e)

        try:
            r_filter = requests.get(f"{base_url}/leads?lead_status=New&limit=5", headers={"Authorization": f"Bearer {admin_token}"}, timeout=10)
            out["admin_filter_status"] = r_filter.status_code
        except Exception as e:
            out["admin_filter_error"] = str(e)

        # timeline for a sample lead
        try:
            sample_id = None
            if out.get("admin_leads_list_preview") and isinstance(out["admin_leads_list_preview"], list):
                sample_id = out["admin_leads_list_preview"][0].get('id')
            if sample_id:
                r_tl = requests.get(f"{base_url}/timeline/lead/{sample_id}", headers={"Authorization": f"Bearer {admin_token}"}, timeout=10)
                out["admin_timeline_status"] = r_tl.tatus_code
                out["admin_timeline_preview_count"] = len(r_tl.json()) if r_tl.status_code == 200 else None
        except Exception as e:
            out["admin_timeline_error"] = str(e)

        try:
            r_reports = requests.get(f"{base_url}/eod", headers={"Authorization": f"Bearer {admin_token}"}, timeout=10)
            out["admin_reports_status"] = r_reports.status_code
        except Exception as e:
            out["admin_reports_error"] = str(e)

    # sample employee token (use one seeded email)
    emp_token = get_token(base_url, "vaibhav.borge@fundingsathi.in", "vaibhav.emp@01")
    out["employee_token_present"] = bool(emp_token)
    if emp_token:
        try:
            r = requests.get(f"{base_url}/leads?limit=5", headers={"Authorization": f"Bearer {emp_token}"}, timeout=10)
            out["employee_leads_list_status"] = r.status_code
            out["employee_leads_list_preview"] = r.json() if r.status_code == 200 else r.text[:400]
        except Exception as e:
            out["employee_leads_list_error"] = str(e)

    # Timeline / timestamp preservation checks for random sample of leads
    import random
    timeline_checks = {}
    with database.engine.connect() as conn:
        ids = [r[0] for r in conn.execute(text('SELECT id FROM leads ORDER BY RANDOM() LIMIT 20'))]
        for lid in ids:
            events = [dict(row) for row in conn.execute(text('SELECT id, event_type, description, created_by, created_at FROM timeline_events WHERE lead_id = :lid ORDER BY created_at'), {'lid': lid}).mappings()]
            raws = [dict(row) for row in conn.execute(text('SELECT id, date_of_entry, data FROM lead_import_raw WHERE lead_id = :lid ORDER BY date_of_entry'), {'lid': lid}).mappings()]
            lead_row = conn.execute(text('SELECT id, created_at, updated_at FROM leads WHERE id = :lid'), {'lid': lid}).mappings().first()
            raw_dates = set()
            for r in raws:
                if r.get('date_of_entry'):
                    raw_dates.add(str(r['date_of_entry']).split('+')[0])
                # also inspect JSON payload for potential dates
                try:
                    d = json.loads(r.get('data') or '{}')
                    for v in d.values():
                        if isinstance(v, str) and v[:4].isdigit():
                            raw_dates.add(v.split('+')[0])
                except Exception:
                    pass

            event_date_matches = []
            for ev in events:
                created_at = str(ev.get('created_at'))
                matched = any(created_at.startswith(rd) or rd.startswith(created_at[:10]) for rd in raw_dates) or (lead_row and created_at.startswith(str(lead_row['created_at'])[:10]))
                event_date_matches.append({'event_id': str(ev.get('id')), 'event_type': ev.get('event_type'), 'created_at': created_at, 'matched_in_raw': matched})

            timeline_checks[str(lid)] = {
                'events_count': len(events),
                'raw_rows_count': len(raws),
                'lead_created_at': str(lead_row['created_at']) if lead_row else None,
                'events': event_date_matches,
            }

    out['timeline_checks_sample'] = timeline_checks

    # Performance probes (simple timings)
    perf = {}
    def time_request(method, url, headers=None):
        t0 = time.time()
        try:
            r = requests.request(method, url, headers=headers, timeout=30)
            return (time.time() - t0, r.status_code, r)
        except Exception as e:
            return (time.time() - t0, 0, str(e))

    if admin_token:
        headers = {"Authorization": f"Bearer {admin_token}"}
        elapsed, status, resp = time_request('GET', f"{base_url}/leads?limit=25", headers)
        perf['admin_leads_list'] = {'elapsed_s': elapsed, 'status': status}
        if isinstance(resp, requests.Response):
            perf['admin_leads_list_count'] = len(resp.json()) if resp.status_code == 200 else None
        elapsed, status, resp = time_request('GET', f"{base_url}/leads?search=Vaibhav&limit=25", headers)
        perf['admin_search'] = {'elapsed_s': elapsed, 'status': status}
        if sample_id:
            elapsed, status, resp = time_request('GET', f"{base_url}/timeline/lead/{sample_id}", headers)
            perf['admin_timeline_sample'] = {'elapsed_s': elapsed, 'status': status}

    if emp_token:
        headers = {"Authorization": f"Bearer {emp_token}"}
        elapsed, status, resp = time_request('GET', f"{base_url}/leads?limit=25", headers)
        perf['emp_leads_list'] = {'elapsed_s': elapsed, 'status': status}

    out['performance'] = perf

    # write report
    report_path = Path(__file__).resolve().parents[1] / "reports" / "post_import_verification.json"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, default=str)

    print("Verification report written to", report_path)


if __name__ == "__main__":
    main()
