from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from sqlalchemy.orm import Session
from datetime import datetime

from app.database import get_db
from app.models.lender import Lender
from app.models.lead import Lead
from app.models.lender_case import LenderCase
from lender_models import LeadLenderRequirements, LeadLenderMapping, LenderRecommendation
from app.services.lender_case_service import create_lender_case

try:
    from app.services.forecast_service import ForecastCalculationEngine
    FORECAST_AVAILABLE = True
except ImportError:
    FORECAST_AVAILABLE = False

router = APIRouter(prefix='/api')


class LenderIn(BaseModel):
    name: str
    slug: Optional[str] = None
    logo: Optional[str] = None
    min_turnover: Optional[float] = None
    max_loan: Optional[float] = None
    min_cibil: Optional[int] = None
    roi: Optional[str] = None
    products: Optional[List[str]] = None
    eligible_types: Optional[List[str]] = None
    ticket_min: Optional[float] = 0
    ticket_max: Optional[float] = 0
    min_vintage: Optional[int] = 0
    min_dscr: Optional[float] = None
    requires_atnw_positive: Optional[bool] = False
    requires_owned_property: Optional[bool] = False
    processing_fee: Optional[float] = 0
    foreclosure_charges: Optional[float] = 0
    hidden_charges: Optional[float] = 0
    security_requirement: Optional[str] = None
    property_requirement: Optional[str] = None
    gst: Optional[str] = None
    pan: Optional[str] = None
    cin: Optional[str] = None
    priority_score: Optional[int] = 0
    sla: Optional[str] = None
    average_approval_days: Optional[int] = 0
    average_disbursement_days: Optional[int] = 0
    historical_approval_rate: Optional[float] = 0
    historical_rejection_rate: Optional[float] = 0
    active_status: Optional[bool] = True


class RequirementIn(BaseModel):
    business_type: str
    vintage_years: int
    annual_turnover_cr: float
    cibil: int
    atnw_positive: bool
    dscr: Optional[float] = None
    loan_amount_lakhs: float
    loan_tenure_days: int
    owned_property: bool
    sector: Optional[str] = None
    product_type: Optional[str] = None


def _coerce_list(value: Any) -> List[str]:
    if not value:
        return []
    if isinstance(value, list):
        return [str(item) for item in value if str(item).strip()]
    if isinstance(value, str):
        return [part.strip() for part in value.split(',') if part.strip()]
    return [str(value)]


def _merge_lender_fields(existing: Lender, payload: Dict[str, Any]) -> None:
    for key, value in payload.items():
        if key in {'products', 'eligible_types'}:
            setattr(existing, key, _coerce_list(value))
        else:
            setattr(existing, key, value)


@router.post('/lenders/sync')
def sync_lenders(payload: List[LenderIn], session: Session = Depends(get_db)):
    created = 0
    updated = 0
    for item in payload:
        values = item.dict(exclude_unset=True)
        exists = session.query(Lender).filter_by(name=values.get('name')).first()
        if exists:
            _merge_lender_fields(exists, values)
            updated += 1
        else:
            obj = Lender(**values)
            session.add(obj)
            created += 1
    session.commit()
    return {'status': 'ok', 'created': created, 'updated': updated}


@router.post('/lenders/import-excel')
def import_lenders_from_excel(file: UploadFile = File(...), session: Session = Depends(get_db)):
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(status_code=400, detail='Only .xlsx files are supported')
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f'Excel import requires openpyxl: {exc}') from exc

    import io
    workbook = load_workbook(io.BytesIO(file.file.read()), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {'status': 'ok', 'total_rows': 0, 'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}
    headers = [str(col).strip() if col is not None else '' for col in rows[0]]
    created = 0
    updated = 0
    skipped = 0
    errors = []

    for row_index, row in enumerate(rows[1:], start=2):
        if not any(value is not None and str(value).strip() for value in row):
            skipped += 1
            continue
        record = dict(zip(headers, row))
        name = None
        for key in ('name', 'lender_name', 'lender', 'bank_name'):
            value = record.get(key)
            if value not in (None, ''):
                name = str(value)
                break
        if not name:
            skipped += 1
            continue
        payload = {
            'name': name,
            'slug': str(record.get('slug') or name).lower().replace(' ', '-'),
            'logo': record.get('logo') or None,
            'min_turnover': float(record.get('min_turnover') or 0) if str(record.get('min_turnover') or '').replace('.', '', 1).isdigit() else None,
            'min_cibil': int(record.get('min_cibil') or 0) if str(record.get('min_cibil') or '').isdigit() else None,
            'roi': record.get('roi') or None,
            'products': _coerce_list(record.get('products') or record.get('product')),
            'eligible_types': _coerce_list(record.get('eligible_types') or record.get('business_types')),
            'ticket_min': float(record.get('ticket_min') or 0) if str(record.get('ticket_min') or '').replace('.', '', 1).isdigit() else 0,
            'ticket_max': float(record.get('ticket_max') or 0) if str(record.get('ticket_max') or '').replace('.', '', 1).isdigit() else 0,
            'min_vintage': int(record.get('min_vintage') or 0) if str(record.get('min_vintage') or '').isdigit() else 0,
            'processing_fee': float(record.get('processing_fee') or 0) if str(record.get('processing_fee') or '').replace('.', '', 1).isdigit() else 0,
            'average_approval_days': int(record.get('average_approval_days') or 0) if str(record.get('average_approval_days') or '').isdigit() else 0,
            'average_disbursement_days': int(record.get('average_disbursement_days') or 0) if str(record.get('average_disbursement_days') or '').isdigit() else 0,
            'historical_approval_rate': float(record.get('historical_approval_rate') or 0) if str(record.get('historical_approval_rate') or '').replace('.', '', 1).replace('%', '', 1).isdigit() else 0,
            'active_status': True,
        }
        try:
            exists = session.query(Lender).filter_by(name=name).first()
            if exists:
                _merge_lender_fields(exists, payload)
                updated += 1
            else:
                session.add(Lender(**payload))
                created += 1
        except Exception as exc:
            errors.append({'row': row_index, 'name': name, 'error': str(exc)})
    session.commit()
    return {'status': 'ok', 'total_rows': len(rows) - 1, 'created': created, 'updated': updated, 'skipped': skipped, 'errors': errors}


@router.post('/leads/{lead_id}/capture-lender-requirements')
def capture_requirements(lead_id: str, payload: RequirementIn, session: Session = Depends(get_db)):
    obj = LeadLenderRequirements(lead_id=lead_id, data=payload.dict())
    session.add(obj)
    session.commit()
    return {'status': 'ok', 'id': obj.id}


def _get_requirement_payload(req: Dict[str, Any]) -> Dict[str, Any]:
    return {
        'business_type': req.get('business_type') or 'manufacturing',
        'vintage_years': int(req.get('vintage_years') or 0),
        'annual_turnover_cr': float(req.get('annual_turnover_cr') or 0),
        'cibil': int(req.get('cibil') or 0),
        'atnw_positive': bool(req.get('atnw_positive', False)),
        'dscr': float(req.get('dscr') or 0) if req.get('dscr') not in (None, '') else None,
        'loan_amount_lakhs': float(req.get('loan_amount_lakhs') or 0),
        'loan_tenure_days': int(req.get('loan_tenure_days') or 0),
        'owned_property': bool(req.get('owned_property', False)),
        'sector': req.get('sector') or None,
        'product_type': str(req.get('product_type') or '').strip() or None,
    }


def score_lender_against_req(lender: Lender, req: Dict[str, Any]) -> Dict[str, Any]:
    payload = _get_requirement_payload(req)
    eligibility_score = 0.0
    reasons = []
    loan_amount = float(payload.get('loan_amount_lakhs', 0) or 0) * 100000

    if payload['vintage_years'] >= (lender.min_vintage or 0):
        eligibility_score += 20
        reasons.append('Vintage eligible')
    else:
        reasons.append('Vintage below minimum')

    turnover = float(payload.get('annual_turnover_cr', 0) or 0)
    min_turnover = getattr(lender, 'min_turnover', None)
    turnover_met = False
    if min_turnover is None:
        turnover_met = True
    elif min_turnover >= 10000000:
        turnover_rupees = turnover * 10000000
        turnover_met = turnover_rupees >= min_turnover
    else:
        turnover_met = turnover >= (min_turnover / 10000000)
    if turnover_met:
        eligibility_score += 25
        reasons.append('Turnover eligible')
    else:
        reasons.append('Turnover below minimum')

    if lender.ticket_min and lender.ticket_max:
        if lender.ticket_min <= loan_amount <= lender.ticket_max:
            eligibility_score += 20
            reasons.append('Loan amount fits ticket')
        else:
            reasons.append('Loan amount outside ticket')
    else:
        eligibility_score += 20
        reasons.append('Ticket size not constrained')

    allowed_types = list(getattr(lender, 'eligible_types', []) or [])
    if not allowed_types or payload.get('business_type') in allowed_types:
        eligibility_score += 10
        reasons.append('Industry fit')
    else:
        reasons.append('Industry mismatch')

    product_type = (payload.get('product_type') or '').strip().lower()
    if product_type:
        lender_products_raw = getattr(lender, 'products', []) or []
        if isinstance(lender_products_raw, str):
            lender_products = [p.strip().lower() for p in lender_products_raw.split(',') if p.strip()]
        else:
            lender_products = [str(p).strip().lower() for p in lender_products_raw if str(p).strip()]

        product_match = any(
            product_type in lender_product or lender_product in product_type
            for lender_product in lender_products
        )
        if product_match:
            eligibility_score += 20
            reasons.append('Product match')
        else:
            reasons.append('Product mismatch')
    else:
        reasons.append('Product not provided')

    min_cibil = lender.min_cibil or 650
    if payload['cibil'] >= min_cibil:
        eligibility_score += 25
        reasons.append('CIBIL exceeds threshold')
    else:
        reasons.append('CIBIL below threshold')

    min_dscr = lender.min_dscr
    dscr = payload.get('dscr')
    if min_dscr and dscr is not None and dscr >= min_dscr:
        eligibility_score += 10
        reasons.append('DSCR meets threshold')
    elif not min_dscr:
        eligibility_score += 10
        reasons.append('DSCR not required')
    else:
        reasons.append('DSCR below threshold')

    if lender.requires_atnw_positive and payload.get('atnw_positive'):
        eligibility_score += 5
        reasons.append('ATNW positive')
    elif lender.requires_atnw_positive:
        reasons.append('ATNW missing')
    else:
        eligibility_score += 5
        reasons.append('ATNW not required')

    if lender.requires_owned_property and payload.get('owned_property'):
        eligibility_score += 5
        reasons.append('Property confirmed')
    elif lender.requires_owned_property:
        reasons.append('Property missing')
    else:
        eligibility_score += 5
        reasons.append('Property not required')

    eligibility_pct = round(min(100.0, eligibility_score), 1)
    risk_penalty = 0.0
    if payload['cibil'] < 650:
        risk_penalty += 15
    if turnover < (lender.min_turnover or 0):
        risk_penalty += 20
    if loan_amount > 50000000:
        risk_penalty += 10
    risk_pct = round(max(0.0, 100.0 - risk_penalty), 1)

    approval_pct = round((getattr(lender, 'historical_approval_rate', 0) or 0) * 100, 1)
    match_score = round(min(100.0, (0.6 * eligibility_pct) + (0.15 * risk_pct) + (0.25 * approval_pct)), 1)
    approval_probability = round(min(0.99, max(0.01, ((getattr(lender, 'historical_approval_rate', 0) or 0) + (match_score / 100.0)) / 2.0)), 2)

    return {
        'score': round(match_score, 1),
        'match_score': round(match_score, 1),
        'eligibility_pct': eligibility_pct,
        'risk_pct': risk_pct,
        'approval_pct': approval_pct,
        'approval_probability': approval_probability,
        'rule_score': round(eligibility_pct, 1),
        'ai_score': round((eligibility_pct + risk_pct) / 2.0, 1),
        'historical_score': round(approval_pct, 1),
        'reasons': reasons,
        'roi': lender.roi,
        'lender_id': lender.id,
        'lender_name': lender.name,
        'expected_roi': lender.roi,
        'expected_processing_time': getattr(lender, 'average_approval_days', 0) or 3,
        'approval_days': getattr(lender, 'average_approval_days', 0) or 3,
        'processing_fee': getattr(lender, 'processing_fee', 0) or 0,
        'products': getattr(lender, 'products', []) or [],
    }


def build_lender_recommendation(lender: Lender, req: Dict[str, Any]) -> Dict[str, Any]:
    scored = score_lender_against_req(lender, req)
    parts = []
    if scored['eligibility_pct'] >= 85:
        parts.append('Turnover and vintage eligible')
    if scored['approval_probability'] >= 0.8:
        parts.append('Historical approval')
    if scored['match_score'] >= 80:
        parts.append('Strong product fit')
    if scored['products']:
        parts.append(f"Products: {', '.join(scored['products'][:3])}")
    scored['reason'] = ' · '.join(parts) if parts else 'Meets core eligibility checks'
    scored['recommendation_rank'] = 1
    return scored


@router.get('/leads/{lead_id}/recommended-lenders')
def recommended_lenders(lead_id: str, session: Session = Depends(get_db)):
    requirement_entry = session.query(LeadLenderRequirements).filter_by(lead_id=lead_id).order_by(LeadLenderRequirements.created_at.desc()).first()
    if not requirement_entry:
        raise HTTPException(status_code=404, detail='No requirements captured for lead')
    payload = requirement_entry.data
    lenders = session.query(Lender).filter_by(active_status=True).all()
    scored = []
    for lender in lenders:
        result = build_lender_recommendation(lender, payload)
        scored.append(result)
        session.add(LenderRecommendation(lead_id=lead_id, lender_id=lender.id, score=result['match_score'], details=result))
    session.commit()
    scored_sorted = sorted(scored, key=lambda x: x['match_score'], reverse=True)
    return {'results': scored_sorted}


@router.get('/leads/{lead_id}/recommendations')
def get_recommendations(lead_id: str, session: Session = Depends(get_db)):
    return recommended_lenders(lead_id, session=session)


def complete_lender_workflow(session: Session, lead_id: str, lender_id: int, selected_lender_name: Optional[str], payload: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    lead = session.query(Lead).filter(Lead.id == int(lead_id)).first()
    if not lead:
        raise HTTPException(status_code=404, detail='Lead not found')

    requirement_entry = session.query(LeadLenderRequirements).filter_by(lead_id=lead_id).order_by(LeadLenderRequirements.created_at.desc()).first()
    if not requirement_entry:
        raise HTTPException(status_code=404, detail='No requirements captured for lead')

    payload = payload or {}
    mapping = LeadLenderMapping(lead_id=lead_id, lender_id=lender_id, status='pending', note=payload.get('note') or 'Workflow initiated')
    session.add(mapping)
    session.commit()
    session.refresh(mapping)

    case_payload = {
        'application_id': f'LC-{lead_id}-{lender_id}-{datetime.utcnow().strftime("%Y%m%d%H%M%S")}',
        'lead_id': int(lead_id),
        'parent_lead_id': int(lead_id),
        'lead_company': lead.company_name or lead.lead_name,
        'lender_name': selected_lender_name or f'Lender {lender_id}',
        'product_type': requirement_entry.data.get('business_type') if requirement_entry.data else None,
        'applied_loan_amount': requirement_entry.data.get('loan_amount_lakhs') if requirement_entry.data else 0,
        'application_status': 'Login Initiated',
        'remarks': payload.get('note') or 'Created from lender workflow',
    }
    lender_case = create_lender_case(session, case_payload, creator=None, created_by_name='System')

    forecast_engine = ForecastCalculationEngine(session)
    snapshot = forecast_engine.create_forecast_snapshot(
        lead_id=int(lead_id),
        lead=lead,
        lender_case=lender_case,
        current_stage='Login with Lender',
        loan_amount=case_payload['applied_loan_amount'] * 100000,
        lender_id=str(lender_id),
        product_id=None,
        business_vertical_id=None,
        rm_id=str(lead.assigned_to) if lead.assigned_to else None,
        rm_name=None,
    )
    forecast_engine.record_audit_trail(
        lead_id=int(lead_id),
        change_type='lender_workflow',
        field_name='lead_stage',
        previous_value=lead.lead_status,
        new_value='Bank Selected',
        changed_by=None,
        changed_by_name='System',
        reason='Lender workflow completed',
        revenue_impact=snapshot.expected_revenue or 0,
        weighted_revenue_impact=snapshot.weighted_revenue or 0,
    )

    lead.lead_status = 'Bank Selected'
    lead.pipeline_stage = 'Login with Lender'
    lead.updated_at = datetime.utcnow()
    session.add(lead)
    session.commit()
    session.refresh(lead)

    return {
        'status': 'ok',
        'portal': f'https://lender-portal.example.com/los/{lender_id}?lead={lead_id}',
        'mapping_id': mapping.id,
        'lender_case_id': lender_case.id,
        'forecast_id': str(snapshot.id),
        'lead_status': lead.lead_status,
        'workflow_complete': True,
    }


@router.post('/leads/{lead_id}/select-lender/{lender_id}')
def select_lender(lead_id: str, lender_id: int, session: Session = Depends(get_db), payload: Optional[Dict[str, Any]] = None):
    return complete_lender_workflow(session=session, lead_id=lead_id, lender_id=lender_id, selected_lender_name=None, payload=payload)


@router.post('/leads/{lead_id}/mapping/{mapping_id}/status')
def update_mapping_status(lead_id: str, mapping_id: int, payload: Dict[str, Any], session: Session = Depends(get_db)):
    mapping = session.query(LeadLenderMapping).filter_by(id=mapping_id, lead_id=lead_id).first()
    if not mapping:
        raise HTTPException(status_code=404, detail='Mapping not found')
    status = payload.get('status')
    note = payload.get('note')
    if status:
        mapping.status = status
    if note:
        mapping.note = (mapping.note or '') + '\n' + str(note)
    session.add(mapping)
    session.commit()
    return {'status': 'ok', 'mapping_id': mapping.id, 'new_status': mapping.status}


@router.get('/leads/{lead_id}/mapping/{mapping_id}')
def get_mapping(lead_id: str, mapping_id: int, session: Session = Depends(get_db)):
    mapping = session.query(LeadLenderMapping).filter_by(id=mapping_id, lead_id=lead_id).first()
    if not mapping:
        raise HTTPException(status_code=404, detail='Mapping not found')
    return {'id': mapping.id, 'lead_id': mapping.lead_id, 'lender_id': mapping.lender_id, 'status': mapping.status, 'note': mapping.note}


@router.post('/lenders/run-matching')
def run_matching(lead_id: Optional[str] = None, session: Session = Depends(get_db)):
    if not lead_id:
        return {'status': 'ok', 'results': []}
    requirement_entry = session.query(LeadLenderRequirements).filter_by(lead_id=lead_id).order_by(LeadLenderRequirements.created_at.desc()).first()
    if not requirement_entry:
        raise HTTPException(status_code=404, detail='No requirements captured for lead')
    results = []
    for lender in session.query(Lender).filter_by(active_status=True).all():
        result = build_lender_recommendation(lender, requirement_entry.data)
        results.append(result)
    return {'lead_id': lead_id, 'results': sorted(results, key=lambda item: item['match_score'], reverse=True)}


@router.get('/lender-dashboard')
def lender_dashboard(session: Session = Depends(get_db)):
    total_lenders = session.query(Lender).count()
    active_lenders = session.query(Lender).filter_by(active_status=True).count()
    pending_mappings = session.query(LeadLenderMapping).filter(LeadLenderMapping.status != 'disbursed').count()
    recommendations = session.query(LenderRecommendation).count()
    return {
        'total_lenders': total_lenders,
        'active_lenders': active_lenders,
        'pending_applications': pending_mappings,
        'recommendations_generated': recommendations,
        'top_lender': session.query(Lender).first().name if session.query(Lender).first() else None,
    }


@router.get('/lender-analytics')
def lender_analytics(session: Session = Depends(get_db)):
    recommendations = session.query(LenderRecommendation).all()
    avg_match = round(sum(item.score or 0 for item in recommendations) / len(recommendations), 1) if recommendations else 0
    return {
        'average_match_pct': avg_match,
        'recommendation_count': len(recommendations),
        'pending_applications': session.query(LeadLenderMapping).filter(LeadLenderMapping.status != 'disbursed').count(),
        'mapped_leads': session.query(LeadLenderMapping).count(),
    }


@router.get('/lender-history/{lead_id}')
def lender_history(lead_id: str, session: Session = Depends(get_db)):
    history = session.query(LenderRecommendation).filter_by(lead_id=lead_id).order_by(LenderRecommendation.created_at.desc()).all()
    return {'lead_id': lead_id, 'history': [{'lender_id': item.lender_id, 'score': item.score, 'details': item.details} for item in history]}


@router.post('/lender-status-webhook')
def lender_status_webhook(payload: Dict[str, Any], session: Session = Depends(get_db)):
    mapping_id = payload.get('mapping_id')
    if not mapping_id:
        raise HTTPException(status_code=400, detail='mapping_id is required')
    mapping = session.query(LeadLenderMapping).filter_by(id=int(mapping_id)).first()
    if not mapping:
        raise HTTPException(status_code=404, detail='Mapping not found')
    mapping.status = payload.get('status', mapping.status)
    mapping.note = payload.get('note', mapping.note)
    session.add(mapping)
    session.commit()
    return {'status': 'ok', 'mapping_id': mapping.id, 'new_status': mapping.status}


@router.get('/lenders')
def list_lenders(session: Session = Depends(get_db)):
    lenders = session.query(Lender).all()
    return {'lenders': [{
        'id': l.id,
        'name': l.name,
        'roi': l.roi,
        'ticket_min': l.ticket_min,
        'ticket_max': l.ticket_max,
        'historical_approval_rate': getattr(l, 'historical_approval_rate', 0),
        'processing_fee': getattr(l, 'processing_fee', 0),
    } for l in lenders]}


@router.get('/lenders/{lender_id}')
def get_lender(lender_id: int, session: Session = Depends(get_db)):
    l = session.query(Lender).filter_by(id=lender_id).first()
    if not l:
        raise HTTPException(status_code=404, detail='Lender not found')
    return {
        'id': l.id,
        'name': l.name,
        'logo': l.logo,
        'roi': l.roi,
        'products': l.products,
        'eligible_types': l.eligible_types,
        'processing_fee': getattr(l, 'processing_fee', 0),
        'average_approval_days': getattr(l, 'average_approval_days', 0),
        'historical_approval_rate': getattr(l, 'historical_approval_rate', 0),
        'extra': l.extra,
    }
